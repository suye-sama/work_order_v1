"""
Agent API 路由 — 前端/后端调用 Agent 智能服务的入口
"""
import io
from datetime import datetime as dt

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from app.database import get_db
from app.schemas.common import ApiResponse
from app.middleware.auth import get_current_user
from app.services import agent_service, ticket_service, customer_service

router = APIRouter(prefix="/api/v1/agent", tags=["Agent智能"])


@router.post("/log-parse", response_model=ApiResponse)
async def agent_log_parse(
    ticket_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """调用 Agent 1: 解析工单的原始日志"""
    # 获取工单的 raw_log
    ticket = ticket_service.get_ticket(db, ticket_id)
    if not ticket:
        raise HTTPException(404, "工单不存在")
    if ticket.status != 2:
        raise HTTPException(400, "仅处理中(2)状态的工单可以解析日志")
    if not ticket.raw_log:
        raise HTTPException(400, "工单没有操作日志，请先追加日志")

    result = await agent_service.parse_log(
        log_text=ticket.raw_log,
        ticket_id=ticket_id,
    )

    if result.get("error"):
        raise HTTPException(500, f"Agent 解析失败: {result['error']}")

    # 将时间线以 Agent 生成的方式追加到工单时间线
    for entry in result.get("timeline", [])[:20]:  # 最多 20 条
        from app.models.timeline import TicketTimeline
        from datetime import datetime

        tl = TicketTimeline(
            ticket_id=ticket_id,
            node_time=datetime.now(),
            node_type="Agent解析",
            title=f"[{entry.get('phase', '')}] {entry.get('operation', '')}",
            content=f"命令: {entry.get('command', '')}\n结果: {entry.get('result', '')}",
            operator=str(user.id),
            ai_generated=True,
        )
        db.add(tl)
    db.commit()

    return ApiResponse(code=200, message="日志解析完成", data=result)


@router.post("/generate", response_model=ApiResponse)
async def agent_generate_ticket(
    ticket_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """调用 Agent 2: AI 生成工单内容"""
    ticket = ticket_service.get_ticket(db, ticket_id)
    if not ticket:
        raise HTTPException(404, "工单不存在")
    if ticket.status != 2:
        raise HTTPException(400, "仅处理中(2)状态的工单可以生成工单内容")

    # 准备客户信息
    customer_info = ""
    if ticket.customer:
        customer_info = (
            f"{ticket.customer.name} / "
            f"版本:{ticket.customer.product_version or '未知'} / "
            f"部署:{ticket.customer.deploy_type or '未知'} / "
            f"OS:{ticket.customer.os or '未知'}"
        )

    # 获取时间线数据
    timeline = [t.to_dict() for t in (ticket.timeline or [])]

    result = await agent_service.generate_ticket(
        ticket_id=ticket_id,
        title=ticket.title or "",
        description=ticket.description or "",
        customer_info=customer_info,
        timeline_json=timeline,
        raw_log=ticket.raw_log or "",
        engineer_notes="",
    )

    if not result.get("success"):
        raise HTTPException(500, f"Agent 生成失败: {result.get('error', '未知错误')}")

    # 将结果写入工单
    gen = result.get("result", {})
    if gen:
        fp = gen.get("fault_phenomenon", {})
        rc = gen.get("root_cause", {})
        sol = gen.get("solution", {})
        sm = gen.get("summary_info", {})

        ticket.fault_summary = fp.get("phenomenon", "") if isinstance(fp, dict) else str(fp)
        ticket.root_cause = rc.get("root_cause", "") if isinstance(rc, dict) else str(rc)
        ticket.solution = sol.get("solution", "") if isinstance(sol, dict) else str(sol)
        ticket.ai_summary = sm.get("summary", "") if isinstance(sm, dict) else str(sm)

    # 状态流转到"待确认"
    ticket.status = 3
    db.commit()
    db.refresh(ticket)

    return ApiResponse(
        code=200,
        message="AI 工单生成完成，请审核确认",
        data=ticket.to_dict(include_timeline=True),
    )


@router.post("/extract", response_model=ApiResponse)
async def agent_extract_knowledge(
    ticket_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """调用 Agent 3: 从已完成工单中提取知识"""
    ticket = ticket_service.get_ticket(db, ticket_id)
    if not ticket:
        raise HTTPException(404, "工单不存在")
    if ticket.status not in (4, 5):
        raise HTTPException(400, "仅已完成/已归档的工单可以提取知识")

    result = await agent_service.extract_knowledge(
        ticket_id=ticket_id,
        title=ticket.title or "",
        fault_phenomenon={"phenomenon": ticket.fault_summary or ticket.title or ""},
        root_cause={"root_cause": ticket.root_cause or "", "category": ticket.category or ""},
        solution={"solution": ticket.solution or ""},
        summary_info={"summary": ticket.ai_summary or "", "tags": []},
        category=ticket.category or "",
    )

    if result.get("success"):
        # 记录 knowledge_id 到工单
        ticket.knowledge_id = result.get("knowledge_id")
        db.commit()

    return ApiResponse(
        code=200,
        message="知识提取完成" if result.get("success") else f"提取失败: {result.get('error')}",
        data=result,
    )


@router.post("/analyze", response_model=ApiResponse)
async def agent_analyze(
    ticket_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """调用 Agent 5: 深度故障分析"""
    ticket = ticket_service.get_ticket(db, ticket_id)
    if not ticket:
        raise HTTPException(404, "工单不存在")

    customer_info = ""
    if ticket.customer:
        customer_info = f"{ticket.customer.name} / 版本:{ticket.customer.product_version or '未知'}"

    result = await agent_service.analyze_fault(
        ticket_id=ticket_id,
        title=ticket.title or "",
        fault_phenomenon={"phenomenon": ticket.fault_summary or ticket.title or ""},
        root_cause={"root_cause": ticket.root_cause or "", "category": ticket.category or ""},
        solution={"solution": ticket.solution or ""},
        category=ticket.category or "",
        customer_info=customer_info,
    )

    return ApiResponse(
        code=200,
        message="故障分析完成",
        data=result,
    )


@router.get("/search", response_model=ApiResponse)
async def agent_search(
    q: str = Query(..., description="搜索关键词"),
    top_k: int = Query(5, ge=1, le=50),
    user=Depends(get_current_user),
):
    """调用 Agent 4: 相似工单检索"""
    result = await agent_service.search_similar(q, top_k)
    return ApiResponse(code=200, message="success", data=result)


# ==================== 知识库管理 ====================


@router.get("/knowledge", response_model=ApiResponse)
async def knowledge_list(
    keyword: str = Query("", description="搜索关键词"),
    category: str = Query("", description="分类筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    user=Depends(get_current_user),
):
    """知识库列表 — 分页查询，支持搜索"""
    result = await agent_service.list_knowledge(
        keyword=keyword,
        category=category,
        page=page,
        page_size=page_size,
    )
    if result.get("error") and not result.get("items") and result.get("total", 0) == 0:
        raise HTTPException(500, f"知识库查询失败: {result['error']}")
    return ApiResponse(code=200, message="success", data=result)


@router.get("/knowledge/{knowledge_id}", response_model=ApiResponse)
async def knowledge_detail(
    knowledge_id: str,
    user=Depends(get_current_user),
):
    """知识条目详情"""
    result = await agent_service.get_knowledge_detail(knowledge_id)
    if result.get("error"):
        raise HTTPException(500, f"查询失败: {result['error']}")
    return ApiResponse(code=200, message="success", data=result)


@router.get("/knowledge/export/download")
async def knowledge_export_excel(
    keyword: str = Query("", description="搜索关键词"),
    category: str = Query("", description="分类筛选"),
    user=Depends(get_current_user),
):
    """知识库导出 Excel — 包含四段式文档详情"""
    result = await agent_service.list_knowledge(
        keyword=keyword,
        category=category,
        page=1,
        page_size=10000,
    )
    items = result.get("items", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "知识库"

    # 表头
    headers = [
        "序号", "标题", "分类", "难度", "标签",
        "问题描述", "发生原因", "可能产生的现象", "解决方法参考",
        "操作步骤", "预防措施", "关联工单", "条目ID",
    ]
    ws.append(headers)

    # 表头样式
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, item in enumerate(items, 1):
        ws.append([
            i,
            item.get("title", ""),
            item.get("category", ""),
            item.get("difficulty", ""),
            item.get("tags", ""),
            item.get("problem_description", ""),
            item.get("root_cause", ""),
            item.get("symptoms", ""),
            item.get("solution", ""),
            item.get("steps", ""),
            item.get("prevention", ""),
            f"#{item.get('ticket_id')}" if item.get("ticket_id") else "",
            item.get("id", ""),
        ])

    # 列宽
    col_widths = [6, 22, 10, 8, 18, 40, 40, 40, 40, 35, 30, 12, 14]
    for idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    # 文字换行
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_align

    # 冻结首行
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"knowledge_export_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== Agent 6: 排查方向建议 ====================


@router.post("/suggest/{ticket_id}", response_model=ApiResponse)
async def agent_suggest(
    ticket_id: int,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """调用 Agent 6: 根据工单描述生成排查方向建议，并写入时间线（仅限一次）"""
    ticket = ticket_service.get_ticket(db, ticket_id)
    if not ticket:
        raise HTTPException(404, "工单不存在")
    if ticket.status not in (1, 2):
        raise HTTPException(400, "仅新建(1)或处理中(2)状态的工单可以生成排查建议")

    # 检查是否已生成过建议
    from app.models.timeline import TicketTimeline
    existing = db.query(TicketTimeline).filter(
        TicketTimeline.ticket_id == ticket_id,
        TicketTimeline.node_type == "Agent建议",
        TicketTimeline.ai_generated == True,
    ).first()
    if existing:
        raise HTTPException(400, "该工单已生成过排查建议，请查看时间线")

    # 准备客户信息
    customer_info = ""
    if ticket.customer:
        customer_info = (
            f"{ticket.customer.name} / "
            f"版本:{ticket.customer.product_version or '未知'} / "
            f"部署:{ticket.customer.deploy_type or '未知'} / "
            f"OS:{ticket.customer.os or '未知'}"
        )

    result = await agent_service.suggest_checks(
        title=ticket.title or "",
        description=ticket.description or "",
        category=ticket.category or "",
        customer_info=customer_info,
    )

    if not result.get("success"):
        raise HTTPException(500, f"建议生成失败: {result.get('error', '未知错误')}")

    # 将建议写入时间线（不写入 raw_log）
    from datetime import datetime as dt_module

    # 构建时间线内容
    content_lines = [
        f"【综合分析】{result.get('brief_analysis', '')}",
        "",
    ]
    causes = result.get("possible_causes", [])
    if causes:
        content_lines.append("【可能原因】")
        for i, c in enumerate(causes, 1):
            content_lines.append(f"  {i}. {c}")
        content_lines.append("")

    checks = result.get("suggested_checks", [])
    if checks:
        content_lines.append("【建议检查项】")
        for i, c in enumerate(checks, 1):
            content_lines.append(f"  {i}. {c.get('direction', '')}")
            content_lines.append(f"     原因: {c.get('why', '')}")
            content_lines.append(f"     命令: {c.get('command', '')}")
            if c.get("expected_if_problem"):
                content_lines.append(f"     预期: {c.get('expected_if_problem', '')}")
        content_lines.append("")

    similar = result.get("similar_cases", [])
    if similar:
        content_lines.append("【参考历史案例】")
        for i, c in enumerate(similar[:3], 1):
            sim_pct = f"{c.get('similarity', 0) * 100:.0f}%"
            content_lines.append(f"  {i}. {c.get('title', '')}（相似度 {sim_pct}）")

    tl = TicketTimeline(
        ticket_id=ticket_id,
        node_time=dt_module.now(),
        node_type="Agent建议",
        title="🤖 AI 排查方向建议",
        content="\n".join(content_lines),
        operator=str(user.id),
        ai_generated=True,
    )
    db.add(tl)
    db.commit()

    return ApiResponse(
        code=200,
        message="排查建议已生成",
        data=result,
    )
